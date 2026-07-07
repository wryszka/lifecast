# Databricks notebook source
# MAGIC %md
# MAGIC # LifeCast — Phase 3: CFO export (always offered)
# MAGIC
# MAGIC Genie and the dashboard replace the *rebuilding* of the board pack — not the
# MAGIC CFO's spreadsheet. This task writes the latest-quarter board pack to the volume
# MAGIC as **Excel + CSV** on every results run. QRT / XBRL templates stay in Excel —
# MAGIC we connect, we don't replace.

# COMMAND ----------

# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

# Default catalog for interactive use (jobs always pass ${var.catalog}).
# Porting to another workspace: change once here or via sed across notebooks.
dbutils.widgets.text("catalog", "lr_dev_aws_us_catalog")
CATALOG = dbutils.widgets.get("catalog")
assert CATALOG, "Pass the target catalog via the `catalog` job parameter / widget."

SCHEMA = "lifecast"
FQ = f"`{CATALOG}`.`{SCHEMA}`"
VOLUME_ROOT = f"/Volumes/{CATALOG}/{SCHEMA}/lifecast_files"

# COMMAND ----------

import shutil

import pandas as pd
from pyspark.sql import functions as F

latest = spark.table(f"{FQ}.gld_results_by_product").agg(F.max("reporting_period")).first()[0]

pack = (
    spark.table(f"{FQ}.gld_bel_movement")
    .filter(F.col("reporting_period") == latest)
    .orderBy("product_line")
    .toPandas()
)

out_dir = f"{VOLUME_ROOT}/export/board_pack"
dbutils.fs.mkdirs(out_dir)

csv_path = f"{out_dir}/board_pack_{latest}.csv"
pack.to_csv(csv_path, index=False)

# COMMAND ----------

from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws.title = f"BEL {latest}"
ws["A1"] = f"Bricksurance Life — board pack extract, {latest}"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Source: governed results layer (gld_bel_movement) — synthetic demo data, illustrative only."

headers = ["Product line", "BEL", "BEL prior qtr", "Movement", "Movement %", "Policies"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = Font(bold=True)
for r, row in enumerate(pack.itertuples(index=False), start=5):
    ws.cell(row=r, column=1, value=row.product_line)
    ws.cell(row=r, column=2, value=float(row.bel))
    ws.cell(row=r, column=3, value=float(row.bel_prev) if row.bel_prev is not None else None)
    ws.cell(row=r, column=4, value=float(row.bel_movement) if row.bel_movement is not None else None)
    ws.cell(row=r, column=5, value=float(row.movement_pct) if row.movement_pct is not None else None)
    ws.cell(row=r, column=6, value=int(row.policy_count))
total_row = 5 + len(pack)
ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=total_row, column=2, value=float(pack.bel.sum())).font = Font(bold=True)
ws.cell(row=total_row + 2, column=1,
        value="QRT / XBRL regulatory templates remain in Excel — connect, don't replace.")
for col, width in zip("ABCDEF", [24, 16, 16, 14, 12, 12]):
    ws.column_dimensions[col].width = width

# COMMAND ----------

# MAGIC %md ## Sheet 2 — the rate-risk map (from the engine's own ±100bp runs)

# COMMAND ----------

sens = (spark.table(f"{FQ}.gld_bel_sensitivity")
        .toPandas()
        .pivot_table(index="age_band", columns="term_band", values="dv100_up", aggfunc="sum")
        .sort_index())
TERM_ORDER = ["0–5y", "6–10y", "11–20y", "21y+"]
sens = sens[[t for t in TERM_ORDER if t in sens.columns]]

ws2 = wb.create_sheet("Rate risk")
ws2["A1"] = "BEL change for +100bp, by attained age × outstanding term (£)"
ws2["A1"].font = Font(bold=True, size=12)
ws2.cell(row=3, column=1, value="Age band").font = Font(bold=True)
for c, term in enumerate(sens.columns, start=2):
    ws2.cell(row=3, column=c, value=str(term)).font = Font(bold=True)
for r_i, (age, row) in enumerate(sens.iterrows(), start=4):
    ws2.cell(row=r_i, column=1, value=str(age))
    for c, term in enumerate(sens.columns, start=2):
        v = row[term]
        ws2.cell(row=r_i, column=c, value=None if pd.isna(v) else float(v))
ws2.column_dimensions["A"].width = 12

# COMMAND ----------

# MAGIC %md ## Sheet 3 — provenance. Even the spreadsheet carries its papers.

# COMMAND ----------

aud = (spark.table(f"{FQ}.gld_run_audit")
       .orderBy(F.col("engine_run_at").desc()).limit(1).toPandas())

ws3 = wb.create_sheet("Audit")
ws3["A1"] = "Where these numbers came from (gld_run_audit, latest engine run)"
ws3["A1"].font = Font(bold=True, size=12)
for r_i, (field, value) in enumerate(aud.iloc[0].items(), start=3):
    ws3.cell(row=r_i, column=1, value=str(field)).font = Font(bold=True)
    ws3.cell(row=r_i, column=2, value="" if value is None else str(value))
ws3.column_dimensions["A"].width = 26
ws3.column_dimensions["B"].width = 60

local = "/tmp/board_pack.xlsx"
wb.save(local)
xlsx_path = f"{out_dir}/board_pack_{latest}.xlsx"
shutil.copyfile(local, xlsx_path)

print(f"Board pack exported: {xlsx_path} and {csv_path}")
print(pack[["product_line", "bel", "bel_movement", "movement_pct"]].to_string(index=False))
