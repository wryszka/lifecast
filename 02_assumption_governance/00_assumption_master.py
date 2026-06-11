# Databricks notebook source
# MAGIC %md
# MAGIC # LifeCast — Phase 2: assumption master (tables, functions, seed basis, Excel template)
# MAGIC
# MAGIC Databricks becomes the **master** for mortality / lapse / expense assumptions;
# MAGIC Excel stays where the actuary enters and peer-reviews them.
# MAGIC
# MAGIC - `asm_mortality` / `asm_lapse` / `asm_expense` — versioned by `assumption_set_id`
# MAGIC - `asm_assumption_sets` — the registry: one row per set, with status
# MAGIC   (`DRAFT → PENDING_APPROVAL → APPROVED / REJECTED`, previous basis → `SUPERSEDED`)
# MAGIC - `asm_approval_log` — append-only audit trail of every submit/approve/reject
# MAGIC - UC functions `asm_active_set_id()` + `asm_*_active()` — the single read path
# MAGIC   (notebooks, pipeline, Genie and Excel all resolve the same approved basis)
# MAGIC - Excel entry template (with `DATABRICKS.SQL` formulas) → `lifecast_files/excel/`
# MAGIC
# MAGIC Idempotent: tables `IF NOT EXISTS`; the baseline basis is seeded only when the
# MAGIC registry is empty. The seeded values are **illustrative synthetic data, not a
# MAGIC real basis** — entirely parametric, resembling no published table.

# COMMAND ----------

# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

# Default catalog for interactive use (jobs always pass ${var.catalog}).
# Porting to another workspace: change once here or via sed across notebooks.
dbutils.widgets.text("catalog", "lr_dev_aws_us_catalog")
CATALOG = dbutils.widgets.get("catalog")
assert CATALOG, "Pass the target catalog via the `catalog` job parameter / widget."

SCHEMA = "lifecast"
FQ = f"`{CATALOG}`.`{SCHEMA}`"
VOLUME_ROOT = f"/Volumes/{CATALOG}/{SCHEMA}/lifecast_files"

# COMMAND ----------

# MAGIC %md ## Tables

# COMMAND ----------

spark.sql(f"""
CREATE TABLE IF NOT EXISTS {FQ}.asm_assumption_sets (
    assumption_set_id STRING NOT NULL,
    version INT NOT NULL,
    basis_name STRING NOT NULL,
    status STRING NOT NULL,          -- DRAFT | PENDING_APPROVAL | APPROVED | REJECTED | SUPERSEDED
    source STRING NOT NULL,          -- SEED | EXCEL_ROUNDTRIP
    created_by STRING NOT NULL,
    created_at TIMESTAMP NOT NULL,
    submitted_by STRING,
    submitted_at TIMESTAMP,
    approved_by STRING,
    approved_at TIMESTAMP,
    note STRING
) COMMENT 'Assumption set registry — one row per versioned basis. Databricks is master; Excel is the entry surface.'
""")

spark.sql(f"""
CREATE TABLE IF NOT EXISTS {FQ}.asm_approval_log (
    log_ts TIMESTAMP NOT NULL,
    assumption_set_id STRING NOT NULL,
    action STRING NOT NULL,          -- CREATE | SUBMIT | APPROVE | REJECT | SUPERSEDE
    actor STRING NOT NULL,
    note STRING
) COMMENT 'Append-only audit trail of assumption governance actions — the reproducibility record for SII / IFRS 17 sign-off.'
""")

spark.sql(f"""
CREATE TABLE IF NOT EXISTS {FQ}.asm_mortality (
    assumption_set_id STRING NOT NULL,
    age INT NOT NULL,
    sex STRING NOT NULL,
    smoker_status STRING NOT NULL,
    qx DOUBLE NOT NULL
) COMMENT 'Mortality assumptions (annual qx), versioned by assumption_set_id. Illustrative synthetic basis — not a real table.'
""")

spark.sql(f"""
CREATE TABLE IF NOT EXISTS {FQ}.asm_lapse (
    assumption_set_id STRING NOT NULL,
    policy_year INT NOT NULL,
    lapse_rate DOUBLE NOT NULL
) COMMENT 'Lapse assumptions by policy duration, versioned by assumption_set_id. Illustrative synthetic basis.'
""")

spark.sql(f"""
CREATE TABLE IF NOT EXISTS {FQ}.asm_expense (
    assumption_set_id STRING NOT NULL,
    expense_type STRING NOT NULL,
    value DOUBLE NOT NULL,
    unit STRING NOT NULL
) COMMENT 'Expense assumptions, versioned by assumption_set_id. Illustrative synthetic basis.'
""")

# COMMAND ----------

# MAGIC %md ## UC functions — the single read path for the approved basis

# COMMAND ----------

spark.sql(f"""
CREATE OR REPLACE FUNCTION {FQ}.asm_active_set_id()
RETURNS STRING
COMMENT 'The currently APPROVED assumption set id — the one basis every consumer (pipeline, Genie, Excel) resolves.'
RETURN (SELECT assumption_set_id FROM {FQ}.asm_assumption_sets
        WHERE status = 'APPROVED' ORDER BY approved_at DESC LIMIT 1)
""")

spark.sql(f"""
CREATE OR REPLACE FUNCTION {FQ}.asm_mortality_active()
RETURNS TABLE (age INT, sex STRING, smoker_status STRING, qx DOUBLE)
COMMENT 'Mortality rates of the currently approved basis.'
RETURN SELECT age, sex, smoker_status, qx FROM {FQ}.asm_mortality
       WHERE assumption_set_id = (SELECT assumption_set_id FROM {FQ}.asm_assumption_sets
                                  WHERE status = 'APPROVED' ORDER BY approved_at DESC LIMIT 1)
""")

spark.sql(f"""
CREATE OR REPLACE FUNCTION {FQ}.asm_lapse_active()
RETURNS TABLE (policy_year INT, lapse_rate DOUBLE)
COMMENT 'Lapse rates of the currently approved basis.'
RETURN SELECT policy_year, lapse_rate FROM {FQ}.asm_lapse
       WHERE assumption_set_id = (SELECT assumption_set_id FROM {FQ}.asm_assumption_sets
                                  WHERE status = 'APPROVED' ORDER BY approved_at DESC LIMIT 1)
""")

spark.sql(f"""
CREATE OR REPLACE FUNCTION {FQ}.asm_expense_active()
RETURNS TABLE (expense_type STRING, value DOUBLE, unit STRING)
COMMENT 'Expense assumptions of the currently approved basis.'
RETURN SELECT expense_type, value, unit FROM {FQ}.asm_expense
       WHERE assumption_set_id = (SELECT assumption_set_id FROM {FQ}.asm_assumption_sets
                                  WHERE status = 'APPROVED' ORDER BY approved_at DESC LIMIT 1)
""")

spark.sql(f"""
CREATE OR REPLACE VIEW {FQ}.asm_governance_dashboard
COMMENT 'Assumption sets latest-first with their status and audit fields — the governance destination the Cockpit links to.'
AS SELECT s.*,
          (SELECT COUNT(*) FROM {FQ}.asm_mortality m WHERE m.assumption_set_id = s.assumption_set_id) AS mortality_rows,
          (SELECT COUNT(*) FROM {FQ}.asm_lapse l WHERE l.assumption_set_id = s.assumption_set_id) AS lapse_rows,
          (SELECT COUNT(*) FROM {FQ}.asm_expense e WHERE e.assumption_set_id = s.assumption_set_id) AS expense_rows
   FROM {FQ}.asm_assumption_sets s
   ORDER BY version DESC
""")

print("Tables, functions and governance view in place.")

# COMMAND ----------

# MAGIC %md ## Seed the baseline basis (only if the registry is empty)
# MAGIC With the full maker→checker trail, so the audit log is populated from version 1.

# COMMAND ----------

import datetime
import uuid

import numpy as np
import pandas as pd

if spark.table(f"{FQ}.asm_assumption_sets").count() > 0:
    print("Registry already seeded — skipping.")
else:
    set_id = f"AS_{datetime.date.today():%Y%m%d}_{uuid.uuid4().hex[:6].upper()}"
    user = spark.sql("SELECT current_user()").first()[0]
    now = datetime.datetime.now()

    # Illustrative synthetic mortality: smooth parametric curve (datagen only —
    # not a published table, not a recommendation of how to set a basis).
    ages = np.arange(18, 111)
    rows = []
    for sex, sex_f in [("M", 1.00), ("F", 0.82)]:
        for smoker, smk_f in [("N", 1.00), ("S", 1.85)]:
            qx = (0.00018 + 0.0000022 * np.exp(0.105 * ages)) * sex_f * smk_f
            qx = np.minimum(np.round(qx, 6), 1.0)
            rows += [(set_id, int(a), sex, smoker, float(q)) for a, q in zip(ages, qx)]
    mortality = pd.DataFrame(rows, columns=["assumption_set_id", "age", "sex", "smoker_status", "qx"])

    years = np.arange(1, 41)
    lapse = pd.DataFrame({
        "assumption_set_id": set_id,
        "policy_year": years,
        "lapse_rate": np.round(0.10 * np.exp(-0.18 * (years - 1)) + 0.04, 4),
    })

    expense = pd.DataFrame(
        [
            (set_id, "initial_per_policy", 280.0, "GBP"),
            (set_id, "maintenance_per_policy_pa", 62.0, "GBP"),
            (set_id, "claim_handling_per_claim", 350.0, "GBP"),
            (set_id, "expense_inflation_pa", 0.035, "rate"),
        ],
        columns=["assumption_set_id", "expense_type", "value", "unit"],
    )

    # Explicit schemas — pandas int64 would otherwise clash with the INT columns.
    spark.createDataFrame(
        mortality, "assumption_set_id STRING, age INT, sex STRING, smoker_status STRING, qx DOUBLE"
    ).write.mode("append").saveAsTable(f"{FQ}.asm_mortality")
    spark.createDataFrame(
        lapse, "assumption_set_id STRING, policy_year INT, lapse_rate DOUBLE"
    ).write.mode("append").saveAsTable(f"{FQ}.asm_lapse")
    spark.createDataFrame(
        expense, "assumption_set_id STRING, expense_type STRING, value DOUBLE, unit STRING"
    ).write.mode("append").saveAsTable(f"{FQ}.asm_expense")

    registry_schema = (
        "assumption_set_id STRING, version INT, basis_name STRING, status STRING, source STRING, "
        "created_by STRING, created_at TIMESTAMP, submitted_by STRING, submitted_at TIMESTAMP, "
        "approved_by STRING, approved_at TIMESTAMP, note STRING"
    )
    spark.createDataFrame(
        [(set_id, 1, "Best Estimate baseline", "APPROVED", "SEED",
          user, now, user, now, user, now,
          "Seeded baseline basis (Phase 2 foundation).")],
        registry_schema,
    ).write.mode("append").saveAsTable(f"{FQ}.asm_assumption_sets")

    log_schema = "log_ts TIMESTAMP, assumption_set_id STRING, action STRING, actor STRING, note STRING"
    spark.createDataFrame(
        [
            (now, set_id, "CREATE", user, "Baseline basis created by seed."),
            (now, set_id, "SUBMIT", user, "Baseline submitted for approval."),
            (now, set_id, "APPROVE", user, "Baseline approved as the opening basis."),
        ],
        log_schema,
    ).write.mode("append").saveAsTable(f"{FQ}.asm_approval_log")

    print(f"Seeded baseline basis {set_id} (v1, APPROVED): "
          f"{len(mortality):,} mortality rows, {len(lapse)} lapse rows, {len(expense)} expense rows.")

# COMMAND ----------

# MAGIC %md ## Excel entry template → volume
# MAGIC The first Excel connection point: the actuary keeps entering shocks and loadings
# MAGIC in Excel; the template's `DATABRICKS.SQL` formulas read the approved basis from —
# MAGIC and submit drafts to — the Delta master. Regenerated on every run so the embedded
# MAGIC catalog name always matches this workspace.

# COMMAND ----------

from openpyxl import Workbook

T = f"{CATALOG}.lifecast"
wb = Workbook()

ws = wb.active
ws.title = "README"
for i, line in enumerate(
    [
        "Bricksurance LifeCast — assumption entry template",
        "",
        "About this demo: Bricksurance Life is fictional; all assumptions are synthetic and",
        "illustrative. This template shows the Excel connection point, not a real basis.",
        "",
        "Requires the Databricks Excel add-in (provides the =DATABRICKS.SQL formula).",
        "Excel remains the entry surface; Databricks is the governed master:",
        "  1. 'Current basis' reads the APPROVED basis live from the Delta master.",
        "  2. 'Submit shock' drafts a new basis from your entries and submits it for approval.",
        "  3. Approval happens in Databricks (job lifecast_assumption_approval) — maker and",
        "     checker are separated; every action lands in the asm_approval_log audit trail.",
    ],
    start=1,
):
    ws.cell(row=i, column=1, value=line)

ws2 = wb.create_sheet("Current basis")
ws2["A1"] = "Approved basis id:"
ws2["B1"] = f'=DATABRICKS.SQL("SELECT {T}.asm_active_set_id()")'
ws2["A3"] = "Mortality rates of the approved basis (live read):"
ws2["A4"] = f'=DATABRICKS.SQL("SELECT age, sex, smoker_status, qx FROM {T}.asm_mortality_active() ORDER BY age, sex, smoker_status")'

ws3 = wb.create_sheet("Submit shock")
ws3["A1"] = "New basis name:"
ws3["B1"] = "Best Estimate — smoker loading review"
ws3["A2"] = "New set id (choose one, e.g. AS_YYYYMMDD_REVIEW1):"
ws3["B2"] = "AS_REVIEW1"
ws3["A3"] = "Smoker mortality loading % (applied to qx of smokers):"
ws3["B3"] = 10
ws3["A5"] = "Step 1 — write the draft mortality rows (recalculate this cell):"
ws3["A6"] = (
    f'=DATABRICKS.SQL("INSERT INTO {T}.asm_mortality SELECT \'"&B2&"\', age, sex, smoker_status, '
    f"CASE WHEN smoker_status = 'S' THEN LEAST(ROUND(qx * (1 + \"&B3&\" / 100), 6), 1.0) ELSE qx END "
    f'FROM {T}.asm_mortality_active()")'
)
ws3["A8"] = "Step 2 — register and submit the draft (run these in the SQL editor, or let the"
ws3["A9"] = "lifecast_assumption_entry job do steps 1–2 identically, end to end):"
ws3["A10"] = (
    f"INSERT INTO {T}.asm_lapse SELECT '<set id>', policy_year, lapse_rate FROM {T}.asm_lapse_active(); "
    f"INSERT INTO {T}.asm_expense SELECT '<set id>', expense_type, value, unit FROM {T}.asm_expense_active();"
)
ws3["A11"] = (
    f"INSERT INTO {T}.asm_assumption_sets VALUES ('<set id>', <version>, '<basis name>', 'PENDING_APPROVAL', "
    f"'EXCEL_ROUNDTRIP', current_user(), current_timestamp(), current_user(), current_timestamp(), NULL, NULL, '<note>');"
)
ws3["A12"] = (
    f"INSERT INTO {T}.asm_approval_log VALUES (current_timestamp(), '<set id>', 'SUBMIT', current_user(), '<note>');"
)

excel_dir = f"{VOLUME_ROOT}/excel"
dbutils.fs.mkdirs(excel_dir)
excel_path = f"{excel_dir}/lifecast_assumption_entry.xlsx"
# openpyxl needs seekable writes — save locally, then copy to the volume
# (sequential write; dbutils.fs can't touch /tmp on serverless).
import shutil

local_path = "/tmp/lifecast_assumption_entry.xlsx"
wb.save(local_path)
shutil.copyfile(local_path, excel_path)
print(f"Excel entry template written: {excel_path}")
