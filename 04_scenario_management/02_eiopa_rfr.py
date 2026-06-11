# Databricks notebook source
# MAGIC %md
# MAGIC # LifeCast — Phase 4: EIOPA risk-free rate curve ingestion
# MAGIC
# MAGIC Lands the EIOPA monthly RFR publication (the `RFR_spot_no_VA` tab) into
# MAGIC `esg_rfr_curve` — the curve the illustrative QuantLib generator calibrates to.
# MAGIC **Reused from the actuarial Excel accelerator** (same file format, same parse).
# MAGIC
# MAGIC A sample EIOPA-format workbook ships with the bundle (`sample_data/`); in a client
# MAGIC setting the monthly download lands on the volume the same way. Idempotent: re-runs
# MAGIC replace the same effective date.

# COMMAND ----------

# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

# Default catalog for interactive use (jobs always pass ${var.catalog}).
# Porting to another workspace: change once here or via sed across notebooks.
dbutils.widgets.text("catalog", "lr_dev_aws_us_catalog")
CATALOG = dbutils.widgets.get("catalog")
assert CATALOG, "Pass the target catalog via the `catalog` job parameter / widget."

SCHEMA = "lifecast"
FQ = f"`{CATALOG}`.`{SCHEMA}`"
VOLUME_ROOT = f"/Volumes/{CATALOG}/{SCHEMA}/lifecast_files"
SAMPLE_DIR = "/Workspace/Shared/lifecast/04_scenario_management/sample_data"

# COMMAND ----------

import glob
import os
import re
import shutil
from datetime import date

from openpyxl import load_workbook

# Land the EIOPA file(s) on the volume — the same drop a client's monthly feed makes.
rfr_dir = f"{VOLUME_ROOT}/esg/rfr_inbound"
dbutils.fs.mkdirs(rfr_dir)
for src in glob.glob(f"{SAMPLE_DIR}/EIOPA_RFR_*.xlsx"):
    dst = f"{rfr_dir}/{os.path.basename(src)}"
    if not os.path.exists(dst):
        shutil.copyfile(src, dst)
        print(f"Landed: {dst}")

# COMMAND ----------

spark.sql(f"""
CREATE TABLE IF NOT EXISTS {FQ}.esg_rfr_curve (
    effective_date DATE NOT NULL,
    currency STRING NOT NULL,
    maturity_years INT NOT NULL,
    spot_rate DOUBLE NOT NULL,
    _source_file STRING
) COMMENT 'EIOPA risk-free spot curves (RFR_spot_no_VA), annual compounding. Calibration input for the illustrative scenario generator.'
""")

DATE_RE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")
rows = []
for path in sorted(glob.glob(f"/Volumes/{CATALOG}/{SCHEMA}/lifecast_files/esg/rfr_inbound/EIOPA_RFR_*.xlsx")):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["RFR_spot_no_VA"]
    m = DATE_RE.search(ws["A2"].value or "")
    assert m, f"No reference date in {path}"
    eff = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    ccy_cols = {}
    for col in range(2, ws.max_column + 1):
        ccy = ws.cell(row=5, column=col).value
        if ccy in ("EUR", "GBP", "USD"):
            ccy_cols[ccy] = col
    for r in range(6, 36):  # maturities 1..30
        mat = ws.cell(row=r, column=1).value
        if mat is None:
            continue
        for ccy, col in ccy_cols.items():
            v = ws.cell(row=r, column=col).value
            if v is not None:
                rows.append((eff, ccy, int(mat), float(v), os.path.basename(path)))
    wb.close()

assert rows, "No RFR rows parsed."
curve_df = spark.createDataFrame(
    rows, "effective_date DATE, currency STRING, maturity_years INT, spot_rate DOUBLE, _source_file STRING"
)
for eff in {r[0] for r in rows}:
    spark.sql(f"DELETE FROM {FQ}.esg_rfr_curve WHERE effective_date = '{eff}'")
curve_df.write.mode("append").saveAsTable(f"{FQ}.esg_rfr_curve")

print(f"{len(rows)} curve points loaded into esg_rfr_curve.")
display(spark.sql(f"""
    SELECT effective_date, currency, COUNT(*) points,
           ROUND(MIN(spot_rate),4) min_rate, ROUND(MAX(spot_rate),4) max_rate
    FROM {FQ}.esg_rfr_curve GROUP BY 1, 2 ORDER BY 1, 2
"""))
